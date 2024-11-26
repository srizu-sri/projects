############################################# Library Management System ##############################################################

import openpyxl
from openpyxl import Workbook
from datetime import date
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

# Dictionaries for genres
fiction = {"To Kill a Mockingbird": ["Harper Lee", 10],
"The Great Gatsby": ["F. Scott Fitzgerald", 10],
"1984": ["George Orwell", 10],
"The Catcher in the Rye": ["J.D. Salinger", 10],
"Pride and Prejudice": ["Jane Austen", 10],
"The Book Thief":["Markus Zusak", 10],
"One Hundred Years of Solitude":["Gabriel García Márquez", 10],
"The Road":["Cormac McCarthy", 10],
"Little Women":["Louisa May Alcott", 10],
"Beloved":["Toni Morrison", 10]}

mystery = {"Gone Girl":["Gillian Flynn", 10] ,
"The Girl with the Dragon Tattoo":["Stieg Larsson", 10],
"The Da Vinci Code":["Dan Brown", 10],
"Sherlock Holmes: The Complete Novels and Stories":["Arthur Conan Doyle", 10],
"Big Little Lies":["Liane Moriarty", 10],
"The Silent Patient":["Alex Michaelides", 10],
"In the Woods":["Tana French", 10],
"The Woman in the Window":["A.J. Finn", 10],
"And Then There Were None":["Agatha Christie", 10],
"The Girl on the Train":["Paula Hawkins", 10]}

sci_fi = {"Dune":["Frank Herbert", 10],
"The Hobbit":["J.R.R. Tolkien", 10],
"Harry Potter and the Sorcerer's Stone":["J.K. Rowling", 10],
"Ender's Game":["Orson Scott Card", 10],
"The Name of the Wind":["Patrick Rothfuss", 10],
"The Fellowship of the Ring":["J.R.R. Tolkien", 10],
"A Game of Thrones by George":["R.R. Martin", 10],
"The Left Hand of Darkness":["Ursula K. Le Guin", 10],
"The Martian":["Andy Weir", 10],
"Neuromancer":["William Gibson", 10]}

non_fi = {"Sapiens: A Brief History of Humankind":["Yuval Noah Harari", 10], 
"Educated":["Tara Westover", 10],
"Becoming":["Michelle Obama", 10],
"The Immortal Life of Henrietta Lacks":["Rebecca Skloot", 10],
"Into the Wild":["Jon Krakauer", 10],
"The Wright Brothers":["David McCullough", 10],
"Unbroken":["Laura Hillenbrand", 10],
"Quiet: The Power of Introverts in a World That Can't Stop Talking":["Susan Cain", 10],
"Thinking, Fast and Slow":["Daniel Kahneman", 10],
"A Brief History of Time":["Stephen Hawking", 10]}


genres = {"Fiction": fiction, "Mystery": mystery, "Sci-Fi": sci_fi, "Non-Fiction": non_fi}

# Initialize Excel file
excel_file = "IssuedBooks.xlsx"

# Function to initialize the Excel file if it doesn't exist
def initialize_excel():
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Issued Books"
        ws.append(["Name", "Email", "Book Title", "Date of Issue"])  # Add headers
        wb.save(excel_file)

initialize_excel()

# Function to save issuer's details to Excel
def save_to_excel(name, email, title):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Issued Books"]
    ws.append([name, email, title, date.today().strftime("%Y-%m-%d")])
    wb.save(excel_file)

# ---- Existing Code ----

root = tk.Tk()
root.title("Library Management System")
root.geometry("700x500")

# Create a Notebook widget for tabs
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both")

# Library Tab
library_tab = ttk.Frame(notebook)
notebook.add(library_tab, text="Library")

# # ---- Library Section ----
selected_genre = tk.StringVar(library_tab)
selected_genre.set("Fiction")

# Dropdown for genre selection in Library Tab
genre_label = tk.Label(library_tab, text="Select Genre:")
genre_label.pack(pady=10)
dropdown = tk.OptionMenu(library_tab, selected_genre, *genres.keys())
dropdown.pack(pady=10)

# Treeview to display books
columns = ("Title", "Author", "Availability", "Status")
tree = ttk.Treeview(library_tab, columns=columns, show="headings")
tree.heading("Title", text="Title")
tree.heading("Author", text="Author")
tree.heading("Availability", text="Availability")
tree.heading("Status", text="Status")
tree.column("Title", width=200)
tree.column("Author", width=150)
tree.column("Availability", width=100)
tree.column("Status", width=100)
tree.pack(expand=True, fill="both")

# Function to show books based on selected genre
def show_books():
    # Clear existing items
    for item in tree.get_children():
        tree.delete(item)

    # Display books for the selected genre
    genre_key = selected_genre.get()
    books = genres.get(genre_key, {})

    for title, details in books.items():
        author, availability = details
        status = "Available" if availability > 0 else "Unavailable"
        tree.insert("", "end", values=(title, author, availability, status))

# # Fine Calculator Tab
fine_calculator_tab = ttk.Frame(notebook)
notebook.add(fine_calculator_tab, text="Fine Calculator")


# Button to show books
show_books_button = tk.Button(library_tab, text="Show Books", command=show_books)
show_books_button.pack(pady=10)

# Function to issue a book
def issue_book():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a book to issue.")
        return

    item_values = tree.item(selected_item[0], "values")
    title = item_values[0]
    genre_key = selected_genre.get()
    books = genres.get(genre_key, {})

    if books[title][1] > 0:  # Check availability
        books[title][1] -= 1
        messagebox.showinfo("Issue Book", f"{title} has been issued successfully!")

        # Prompt window for user details
        def save_details():
            name = name_entry.get()
            email = email_entry.get()
            save_to_excel(name, email, title)  # Save details to Excel
            messagebox.showinfo("Details Saved", f"Issued to: {name}, {email}")
            issue_window.destroy()

        issue_window = tk.Toplevel(root)
        issue_window.title("Enter Details")
        issue_window.geometry("300x200")

        tk.Label(issue_window, text="Enter Name:").pack(pady=5)
        name_entry = tk.Entry(issue_window)
        name_entry.pack(pady=5)

        tk.Label(issue_window, text="Enter Email:").pack(pady=5)
        email_entry = tk.Entry(issue_window)
        email_entry.pack(pady=5)

        save_button = tk.Button(issue_window, text="Save Details", command=save_details)
        save_button.pack(pady=10)
    else:
        messagebox.showwarning("Unavailable", "This book is currently unavailable.")

    show_books()

# Add Issue Button
issue_button = tk.Button(library_tab, text="Issue Selected Book", command=issue_book)
issue_button.pack(pady=10)

# Initially show books
show_books()
# ---- Fine Calculator Section ----
# Labels and inputs for Fine Calculator in Fine Calculator Tab
issue_label = tk.Label(fine_calculator_tab, text="Enter Issue Date (yyyy-mm-dd):") #adds label for the user-input field
issue_label.pack(pady=5)
issue_entry = tk.Entry(fine_calculator_tab) #generates an input box for user
issue_entry.pack(pady=5)

return_label = tk.Label(fine_calculator_tab, text="Enter Return Date (yyyy-mm-dd):") #adds label for the user-input field
return_label.pack(pady=5)
return_entry = tk.Entry(fine_calculator_tab) #generates an input box for user
return_entry.pack(pady=5)

fine_label = tk.Label(fine_calculator_tab, text="Fine: ₹0", font=("Arial", 12)) #adds label for the user-input field
fine_label.pack(pady=10)

def calculate_fine():
    try:
        issue_date = date.fromisoformat(issue_entry.get())
        return_date = date.fromisoformat(return_entry.get())
        days_difference = (return_date - issue_date).days
        
        if days_difference > 7:
            fine = 10 * (days_difference - 7)
            fine_label.config(text=f"Fine: ₹{fine}")
        else:
            fine_label.config(text="The book is returned successfully! No fine.")
    except ValueError:
        fine_label.config(text="Invalid date format. Please enter yyyy-mm-dd.")

# Button to calculate fine
calculate_button = tk.Button(fine_calculator_tab, text="Calculate Fine", command=calculate_fine)
calculate_button.pack(pady=10)


root.mainloop()
